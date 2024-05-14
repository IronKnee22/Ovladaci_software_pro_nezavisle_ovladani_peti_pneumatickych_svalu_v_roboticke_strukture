import customtkinter
import time
import openpyxl
from pathlib import Path
import re

import sqlite3

from tkinter import messagebox

from PneumoCVUTFBMI.DeviceLoader import DeviceLoader

# inicializace souboru pro ukladaní měření
workbook = openpyxl.Workbook()
worksheet = workbook.active
Minuly_Sval = 0
Zvoleny_Sval = 0
y = 0
x = 1
pokracovat = True

dl = DeviceLoader()

b1 = dl.getBoard1()
b2 = dl.getBoard2()
b3 = dl.getBoard3()
b4 = dl.getBoard4()
b5 = dl.getBoard5()

# Modes: "System" (standard), "Dark", "Light"
customtkinter.set_appearance_mode("System")
# Themes: "blue" (standard), "green", "dark-blue"
customtkinter.set_default_color_theme("blue")

class adminWindow(customtkinter.CTkToplevel):
    def __init__(self, mainwindow, *args, **kwargs):
        super().__init__(*args, **kwargs)

        width = 480
        height = 240
        
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        screen_x = (screen_width - width) // 2
        screen_y = (screen_height - height) // 2
        self.title("admin")

        self.geometry(f"{width}x{height}+{screen_x}+{screen_y}")

        self.measure_button = customtkinter.CTkButton(self, text="Měření", command= self.mesureWindow) # button for login
        self.measure_button.grid(row =0,column = 1, pady = 10, padx =10)

        self.db_button_sval1 = customtkinter.CTkButton(self, text="Sval1", command=lambda: self.databaseWindow("sval1"))
        self.db_button_sval1.grid(row=1, column=0, pady=10, padx=10)

        self.db_button_sval2 = customtkinter.CTkButton(self, text="Sval2", command=lambda: self.databaseWindow("sval2"))
        self.db_button_sval2.grid(row=1, column=2, pady=10, padx=10)

        self.db_button_sval3 = customtkinter.CTkButton(self, text="Sval3", command=lambda: self.databaseWindow("sval3"))
        self.db_button_sval3.grid(row=3, column=0, pady=10, padx=10)

        self.db_button_sval4 = customtkinter.CTkButton(self, text="Sval4", command=lambda: self.databaseWindow("sval4"))
        self.db_button_sval4.grid(row=3, column=2, pady=10, padx=10)

        self.db_button_sval5 = customtkinter.CTkButton(self, text="Sval5", command=lambda: self.databaseWindow("sval5"))
        self.db_button_sval5.grid(row=2, column=1, pady=10, padx=10)

        self.home_button = customtkinter.CTkButton(self, text="Hlavní stránka", command= self.back2Main) # button for login
        self.home_button.grid(row =4,column = 1, pady = 10, padx =10)

        self.mainwindow = mainwindow
    
    def mesureWindow(self):
        self.toplevel = mesuremenWindow(self,self)
        self.withdraw()

    def databaseWindow(self,sval):
        self.toplevel = databaseWindow(self, sval)
        self.withdraw()

    def back2Main(self):
        # zavře aktualni okno
        self.withdraw()

        # znovu otevře main okno
        self.mainwindow.deiconify()


class mesuremenWindow(customtkinter.CTkToplevel):
    def __init__(self, mainwindow, *args, **kwargs):
        super().__init__(*args, **kwargs)
        
        width = 540
        height = 280

        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        screen_x = (screen_width - width) // 2
        screen_y = (screen_height - height) // 2
        self.title("mesure")

        self.protocol("WM_DELETE_WINDOW", self.execute_after_close)

        self.geometry(f"{width}x{height}+{screen_x}+{screen_y}")

        self.sidebar_frame = customtkinter.CTkFrame(
            self, width=140, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=6, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(4, weight=1)
        self.logo_label = customtkinter.CTkLabel(
            self.sidebar_frame, text="Výběr svalu", font=customtkinter.CTkFont(size=20, weight="bold"))
        self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))

        self.radio_var = customtkinter.IntVar(value=0)

        self.radio_button_1 = customtkinter.CTkRadioButton(
            master=self.sidebar_frame, variable=self.radio_var, value=1, command=self.radiobutton_event, text="Sval 1")
        self.radio_button_1.grid(row=1, column=0)
        self.radio_button_2 = customtkinter.CTkRadioButton(
            master=self.sidebar_frame, variable=self.radio_var, value=2, command=self.radiobutton_event, text="Sval 2")
        self.radio_button_2.grid(row=2, column=0)
        self.radio_button_3 = customtkinter.CTkRadioButton(
            master=self.sidebar_frame, variable=self.radio_var, value=3, command=self.radiobutton_event, text="Sval 3")
        self.radio_button_3.grid(row=3, column=0)
        self.radio_button_4 = customtkinter.CTkRadioButton(
            master=self.sidebar_frame, variable=self.radio_var, value=4, command=self.radiobutton_event, text="Sval 4")
        self.radio_button_4.grid(row=4, column=0)
        self.radio_button_5 = customtkinter.CTkRadioButton(
            master=self.sidebar_frame, variable=self.radio_var, value=5, command=self.radiobutton_event, text="Sval 5")
        self.radio_button_5.grid(row=5, column=0)

        self.main_button_3 = customtkinter.CTkButton(
            master=self.sidebar_frame, fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"), command=self.back2AdminWindow, text="Hl admin")
        self.main_button_3.grid(row=7, column=0, padx=(
            20, 20), pady=(20, 20), sticky="nsew")

        self.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["Light", "Dark", "System"],
                                                                       command=self.change_appearance_mode_event)
        self.appearance_mode_optionemenu.grid(
            row=6, column=0, padx=20, pady=(10, 10))
        self.appearance_mode_optionemenu.set("System")


        self.entrySpeed = customtkinter.CTkEntry(
            self, placeholder_text="Speed")
        self.entrySpeed.configure(validate='focusout', validatecommand=(self.register(self.validate_speed), '%P'))
        self.entrySpeed.grid(row=2, column=1,  padx=(
            20, 0), pady=(20, 20), sticky="nsew")

        self.entrySteps = customtkinter.CTkEntry(
            self, placeholder_text="Steps")
        self.entrySteps.configure(validate='focusout', validatecommand=(self.register(self.validate_steps), '%P'))
        self.entrySteps.grid(row=4, column=1, padx=(
            20, 0), pady=(20, 20), sticky="nsew")

        self.entry = customtkinter.CTkEntry(self, placeholder_text="Hw Value")
        self.entry.configure(validate='focusout', validatecommand=(self.register(self.validate_mbar), '%P'))
        self.entry.grid(row=3, column=1, padx=(20, 0),
                        pady=(20, 20), sticky="nsew")

        self.button_tech_nula = customtkinter.CTkButton(
            self, fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"), text="Technická nula", command=self.technicka_nula)
        self.button_tech_nula.grid(row=2, column=2, padx=(20, 20), pady=(20, 20), sticky="nsew")

        self.main_button_1 = customtkinter.CTkButton(
            self, fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"), command=self.hw_value, text="krok")
        self.main_button_1.grid(row=3, column=2, padx=(
            20, 20), pady=(20, 20), sticky="nsew")
        
        # funkce pomocí které se k nějaká klváese přiřadí určitá klávsa
        self.bind("<Return>", lambda event: self.hw_value())
        self.bind("<KP_Enter>", lambda event: self.hw_value())

        self.main_button_2 = customtkinter.CTkButton(master=self, fg_color="transparent", border_width=2, text_color=(
            "gray10", "#DCE4EE"), command=self.change_pokracovat, text ="pokračovat")
        self.main_button_2.grid(row=4, column=2, padx=(
            20, 20), pady=(20, 20), sticky="nsew")
        self.main_button_2.configure(state='disabled')

        self.mainwindow = mainwindow
        
        
    def technicka_nula(self):
        svaly = {
            0: b1,
            1: b2,
            2: b3,
            3: b4,
            4: b5
            }       

        brokenMuscle = 1

        while True:  
            results = [sval.readA0() for sval in svaly.values()]

            all_in_range = all(600 <= result <= 620 for i, result in enumerate(results) if i != brokenMuscle)

            if all_in_range:
                break  
            
            for i in range(5):
                if i==brokenMuscle:
                    continue
                
                if results[i] >= 620:
                    svaly[i].go_backward(20, 10)
                    
                elif results[i] <= 600:
                    svaly[i].go_forward(20,10)


            time.sleep(1)

    def validate_speed(self, value):
        pattern = r'^\d+$'  # Regulární výraz přijímající pouze čísla
        if re.fullmatch(pattern, value) is None:
            messagebox.showerror("Chybný vstup", "Prosím, zadejte pouze čísla. \nVstup nesmí být prázdný.")
            self.entrySpeed.delete(0, customtkinter.END)            
            return False
        else:
            return True
    
    def validate_steps(self, value):
        pattern = r'^\d+$'  # Regulární výraz přijímající pouze čísla
        if re.fullmatch(pattern, value) is None:
            messagebox.showerror("Chybný vstup", "Prosím, zadejte pouze čísla. \nVstup nesmí být prázdný.")
            self.entrySteps.delete(0, customtkinter.END)            
            return False
        else:
            return True
    
    def validate_mbar(self, value):
        pattern = r'^\d+$'  # Regulární výraz přijímající pouze čísla
        if re.fullmatch(pattern, value) is None:
            messagebox.showerror("Chybný vstup", "Prosím, zadejte pouze čísla. \nVstup nesmí být prázdný.")
            self.entry.delete(0, customtkinter.END)            
            return False
        else:
            return True

    def back2AdminWindow(self):
        # pokud uživatel zavře okno musí dojít k uložení jeho rozpracovaného souboru
        desktop_path = Path.home() / "Desktop"
        workbook.save(desktop_path / f"sval{Minuly_Sval}.xlsx")
        global x
        
        if 'b_objects' in globals() and b_objects is not None:
            if Minuly_Sval in b_objects:
                for i in range(1, x):
                    b_objects[Minuly_Sval].go_backward(Speed, Steps)
                    x=x-1
                    time.sleep(3)
        
        self.withdraw()
        self.mainwindow.deiconify()

    def change_pokracovat(self):
        global pokracovat
        pokracovat = True
        self.main_button_2.configure(state='disabled')

    def hw_value(self):
        speed_value = self.entrySpeed.get()
        steps_value = self.entrySteps.get()
        mbar_value = self.entry.get()

        if Zvoleny_Sval == 0:
            messagebox.showerror("Chybí sval", "Prosím vyber sval")
        
        if self.validate_speed(speed_value) and self.validate_mbar(steps_value) and self.validate_mbar(mbar_value) and Zvoleny_Sval !=0:

            global x
            global y
            global Speed
            global Steps
            global Minuly_Sval
            global pokracovat

            # Vytvoření slovníku pro objekty b1, b2, atd.
            global b_objects
            b_objects = {
                1: b1,
                2: b2,
                3: b3,
                4: b4,
                5: b5
            }

            if Minuly_Sval != Zvoleny_Sval:
                if Minuly_Sval in b_objects:
                    for i in range(1, x):
                        b_objects[Minuly_Sval].go_backward(Speed, Steps)
                        time.sleep(3)

                worksheet.cell(row=1, column=1, value="Počet krokůmotoru")
                worksheet.cell(row=1, column=2, value="Hodnota v mV")
                worksheet.cell(row=1, column=3, value="Hodnota fyzického senzoru")
                worksheet.cell(row=1, column=4, value="Rychlost")
                desktop_path = Path.home() / "Desktop"
                workbook.save(desktop_path / f"sval{Minuly_Sval}.xlsx")

                print("změna svalu")
                x = 1
                y = 0
                Minuly_Sval = Zvoleny_Sval
                pokracovat = False
                self.main_button_2.configure(state='enable')

            if pokracovat == True:
                if Zvoleny_Sval in b_objects:
                    b_object = b_objects[Zvoleny_Sval]

                    b_object.on()
                    Speed = self.entrySpeed.get()
                    Steps = self.entrySteps.get()
                    # Tady se nám počítá kolik udělal motor celkově kroků
                    y = y + int(Steps)
                    x = x + 1
                    b_object.go_forward(Speed, Steps)
                    time.sleep(3)
                    worksheet.cell(row=x, column=1, value=y)  # počet kroků motoru
                    # Zde se načítá ze senzoru může být problém
                    worksheet.cell(row=x, column=2, value=str(b_object.readA0()))
                    # hodnota v mv
                    worksheet.cell(row=x, column=3, value=self.entry.get())
                    # hodnota Rychlosti
                    worksheet.cell(row=x, column=4, value=Speed)
                    Minuly_Sval = Zvoleny_Sval
                    self.entry.delete(0, 10000)
                    print(x)
                else:
                    print("Neplatný Zvoleny_Sval")

    def open_input_dialog_event(self):
        dialog = customtkinter.CTkInputDialog(
            text="Type in a number:", title="CTkInputDialog")
        print("CTkInputDialog:", dialog.get_input())

    def change_appearance_mode_event(self, new_appearance_mode: str):
        customtkinter.set_appearance_mode(new_appearance_mode)

    def getRadiVar(self):
        print("Počáteční hodnota:", self.radio_var.get())

    def radiobutton_event(self):
        global Zvoleny_Sval

        selected_value = self.radio_var.get()

        if 1 <= selected_value <= 5:
            print(f"Sval {selected_value}")
            Zvoleny_Sval = selected_value
        else:
            print("Neznámý sval")

    def execute_after_close(self):

        desktop_path = Path.home() / "Desktop"
        workbook.save(desktop_path / f"sval{Minuly_Sval}.xlsx")
        
        self.destroy()
        svaly = {
            0: b1,
            1: b2,
            2: b3,
            3: b4,
            4: b5
            }       

        brokenMuscle = 1

        while True:  
            results = [sval.readA0() for sval in svaly.values()]

            all_in_range = all(600 <= result <= 620 for i, result in enumerate(results) if i != brokenMuscle)

            if all_in_range:
                break  
            
            for i in range(5):
                if i==brokenMuscle:
                    continue
                
                if results[i] >= 620:
                    svaly[i].go_backward(20, 10)
                elif results[i] <= 600:
                    svaly[i].go_forward(20,10)


            time.sleep(3)

        self.quit()


class databaseWindow(customtkinter.CTkToplevel):
    def __init__(self, mainwindow, sval, *args, **kwargs):
        super().__init__(*args, **kwargs)
        
        width = 1400
        height = 500

        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        screen_x = (screen_width - width) // 2
        screen_y = (screen_height - height) // 2
        self.title(sval)

        self.geometry(f"{width}x{height}+{screen_x}+{screen_y}")

        self.mainwindow = mainwindow
        self.sval = sval
        
        self.start(sval)

    def start(self, sval):
        self.load_table(f"{sval}_mv", 0)
        self.load_table(f"{sval}_mbar", 2)
        self.load_table(f"{sval}_mv2bar", 4)

    def load_table(self, table_name, column):
        global radek
        radek = 1
        
        connection = sqlite3.connect('database.db')
        cursor = connection.cursor()
        cursor.execute(f"SELECT {table_name} FROM {table_name[:5]} WHERE id IN (1)")
        rows = cursor.fetchone()
        
        self.label_auto = customtkinter.CTkLabel(self, text=f"{table_name}: {rows[0]}" )
        self.label_auto.grid(row=0, column=column, padx=10, pady=10)

        self.button_vytvor = customtkinter.CTkButton(self,width= 30, text="nový", command=lambda table_name=table_name: self.vytvorit_mereni(table_name, column))
        self.button_vytvor.grid(row=0, column=column + 1, padx=10)

        self.back2Admin_button = customtkinter.CTkButton(self, text="Hl Admin", command=self.back2AdminWindow, width=20)
        self.back2Admin_button.grid(row=400, column=2, pady = 30)

        self.nacist_z_db(table_name, column)

    def nacist_z_db(self, table_name, column):
        global radek
        connection = sqlite3.connect('database.db')
        cursor = connection.cursor()
        cursor.execute(f"SELECT id, sklon, posun, popis FROM {table_name}")        
        data = cursor.fetchall()
        cursor.close()
        connection.close()

        for row in data:
            id_vzorce, sklon, posun, popis = row
            frame = customtkinter.CTkFrame(self)
            frame.grid(row=id_vzorce, column=column, padx=10, pady=5)

            entry_sklon = customtkinter.CTkEntry(frame, width=80)
            entry_sklon.insert(0, sklon)
            entry_sklon.grid(row=0, column=0, padx=5)

            entry_posun = customtkinter.CTkEntry(frame, width=80)
            entry_posun.insert(0, posun)
            entry_posun.grid(row=0, column=1, padx=5)

            entry_popis = customtkinter.CTkEntry(frame, width=120)
            entry_popis.insert(0, popis)
            entry_popis.grid(row=0, column=2, padx=5)

            button_ulozit = customtkinter.CTkButton(frame, width=20, text="Upravit", command=lambda id=id_vzorce, entry_sklon=entry_sklon, entry_posun=entry_posun, entry_popis=entry_popis: self.ulozit_do_databaze(id, entry_sklon.get(), entry_posun.get(), entry_popis.get(), table_name))
            button_ulozit.grid(row=0, column=3, padx=5)

            radek += 1

    def vytvorit_mereni(self, table_name, column):
        frame = customtkinter.CTkFrame(self)

        # Zjistit počet widgetů ve sloupci a nastavit řádek pro nový rámec
        row = len(self.grid_slaves(column=column)) + 1
        frame.grid(row=row, column=column, padx=10, pady=5)

        self.entry_vyt_posun = customtkinter.CTkEntry(frame, width=80)
        self.entry_vyt_posun.grid(row=0, column=0,padx=5)

        self.entry_vyt_sklon = customtkinter.CTkEntry(frame, width=80)
        self.entry_vyt_sklon.grid(row=0, column=1, padx=5)

        self.entry_vyt_popis = customtkinter.CTkEntry(frame, width=120)
        self.entry_vyt_popis.grid(row=0, column=2, padx=5)

        self.button_ulozit = customtkinter.CTkButton(frame, width=20, text="Uložit", command=lambda table_name=table_name, column=column, frame=frame: self.vlozit_do_databaze(table_name, column, frame))
        self.button_ulozit.grid(row=0, column=3, padx=5)

    def ulozit_do_databaze(self, id_vzorce, sklon, posun, popis, table_name):
        if sklon and posun and popis:
            try:
                connection = sqlite3.connect('database.db')
                cursor = connection.cursor()
                cursor.execute(f"UPDATE {table_name} SET sklon=?, posun=?, popis=? WHERE id=?", (sklon, posun, popis, id_vzorce))
                cursor.execute(f"UPDATE {self.sval} SET {table_name}=? WHERE id=?", (id_vzorce,1))
                connection.commit()
                connection.close()
                print("Změny byly úspěšně uloženy.")

                for widget in self.winfo_children():
                    widget.destroy()
                    
                self.start(self.sval)

            except sqlite3.Error as error:
                print("Chyba při ukládání změn do databáze:", error)
        else:
            print("Prosím, vyplňte všechny pole.")

    def vlozit_do_databaze(self, table_name, column, frame):
        sklon = self.entry_vyt_posun.get()
        posun = self.entry_vyt_sklon.get()
        popis = self.entry_vyt_popis.get()
        if sklon and posun and popis:
            try:
                connection = sqlite3.connect('database.db')
                cursor = connection.cursor()
                cursor.execute(f"INSERT INTO {table_name} (sklon, posun, popis) VALUES (?, ?, ?)", (sklon, posun, popis))
                connection.commit()
                connection.close()
                print("Záznam byl úspěšně uložen do databáze.")

                # Odstranění všech prvků (widgetů) z rámců
                frame.destroy()

                # Aktualizovat zobrazení databáze
                self.nacist_z_db(table_name, column)
            except sqlite3.Error as error:
                print("Chyba při ukládání do databáze:", error)
        else:
            print("Některá pole nejsou vyplněna.")

    def back2AdminWindow(self):
        self.withdraw()
        self.mainwindow.deiconify()

    